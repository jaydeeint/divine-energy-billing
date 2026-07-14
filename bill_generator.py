import os
import sys
from datetime import date

import fitz
import openpyxl
from fpdf import FPDF

MM_TO_PT = 72 / 25.4

if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(__file__)

OUTPUT_DIR = os.path.join(APP_DIR, "bills")
LOGO_PATH = os.path.join(APP_DIR, "assets", "logo.png")
EXCEL_PATH = os.path.join(APP_DIR, "transactions.xlsx")
EXCEL_HEADERS = ["Date", "Bill", "Product", "Quantity", "Price", "Subtotal", "Delivery Charge", "Grand Total"]

PAGE_WIDTH = 80
MARGIN = 4
USABLE_WIDTH = PAGE_WIDTH - (2 * MARGIN)


def get_items():
    items = []
    print("Enter items for the bill. Type 'reset' as the product names to clear all items and start over.\n")
    while True:
        names_raw = input("Product names (separated by ,): ").strip()
        if names_raw.lower() == "reset":
            if items:
                confirm_reset = input(f"Clear all {len(items)} item(s) entered so far? (y/n): ").strip().lower()
                if confirm_reset == "y":
                    items = []
                    print("All items cleared. Starting over.\n")
                else:
                    print("Reset cancelled.\n")
            else:
                print("No items to reset.\n")
            continue

        prices_raw = input("Prices (separated by ,): ").strip()
        quantities_raw = input("Quantities (separated by ,): ").strip()

        names = [n.strip() for n in names_raw.split(",") if n.strip()]
        price_tokens = [p.strip() for p in prices_raw.split(",") if p.strip()]
        quantity_tokens = [q.strip() for q in quantities_raw.split(",") if q.strip()]

        if not names or not (len(names) == len(price_tokens) == len(quantity_tokens)):
            print(
                f"Mismatch: {len(names)} name(s), {len(price_tokens)} price(s), "
                f"{len(quantity_tokens)} quantity(ies). Counts must match. Try again.\n"
            )
            continue

        try:
            prices = [float(p) for p in price_tokens]
            quantities = [int(q) for q in quantity_tokens]
        except ValueError:
            print("Prices must be numbers and quantities must be whole numbers. Try again.\n")
            continue

        batch = [{"name": n, "price": p, "quantity": q} for n, p, q in zip(names, prices, quantities)]

        print("\nConfirm these items:")
        for item in batch:
            subtotal = item["price"] * item["quantity"]
            print(f"  {item['name']}: {item['quantity']} x {item['price']:.2f} = {subtotal:.2f}")
        confirm = input("Add these items? (y/n): ").strip().lower()
        if confirm != "y":
            print("Discarded. Let's re-enter.\n")
            continue

        items.extend(batch)
        print(f"\n{len(batch)} item(s) added. Total items so far: {len(items)}\n")

        generate = input("Generate bill now? (y/n): ").strip().lower()
        if generate == "y":
            break
    return items


def get_delivery_charge():
    while True:
        raw = input("Delivery charge (0 for none): ").strip() or "0"
        try:
            return float(raw)
        except ValueError:
            print("Delivery charge must be a number. Try again.")


def estimate_page_height(items):
    header_height = 46 if os.path.exists(LOGO_PATH) else 20
    items_height = len(items) * 11
    footer_height = 30
    safety_buffer = 20
    return header_height + items_height + footer_height + safety_buffer


class BillPDF(FPDF):
    def header(self):
        if os.path.exists(LOGO_PATH):
            logo_w = 24
            self.image(LOGO_PATH, x=(self.w - logo_w) / 2, y=4, w=logo_w)
            self.set_y(4 + logo_w + 2)
        self.set_font("Helvetica", "B", 12)
        self.cell(0, 6, "RECEIPT", align="C", new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 7)
        self.cell(0, 4, f"Date: {date.today().isoformat()}", align="C", new_x="LMARGIN", new_y="NEXT")
        self.ln(2)


def dashed_separator(pdf):
    y = pdf.get_y() + 1
    pdf.set_dash_pattern(dash=1, gap=1)
    pdf.line(MARGIN, y, PAGE_WIDTH - MARGIN, y)
    pdf.set_dash_pattern()
    pdf.set_y(y + 2)


def build_pdf(items, delivery_charge, output_path):
    pdf = BillPDF()
    pdf.set_margins(MARGIN, MARGIN, MARGIN)
    pdf.set_auto_page_break(False)
    page_height = estimate_page_height(items)
    pdf.add_page(format=(PAGE_WIDTH, page_height))

    items_total = 0.0
    for item in items:
        subtotal = item["price"] * item["quantity"]
        items_total += subtotal

        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(USABLE_WIDTH, 4, item["name"], new_x="LMARGIN", new_y="NEXT")

        pdf.set_font("Helvetica", "", 8)
        pdf.cell(USABLE_WIDTH / 2, 4, f"{item['quantity']} x {item['price']:.2f}")
        pdf.cell(USABLE_WIDTH / 2, 4, f"{subtotal:.2f}", align="R", new_x="LMARGIN", new_y="NEXT")

        dashed_separator(pdf)

    delivery_display = "None" if delivery_charge == 0 else f"{delivery_charge:.2f}"
    grand_total = items_total + delivery_charge

    pdf.set_font("Helvetica", "", 8)
    pdf.cell(USABLE_WIDTH / 2, 5, "Items Total")
    pdf.cell(USABLE_WIDTH / 2, 5, f"{items_total:.2f}", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(USABLE_WIDTH / 2, 5, "Delivery Charge")
    pdf.cell(USABLE_WIDTH / 2, 5, delivery_display, align="R", new_x="LMARGIN", new_y="NEXT")

    dashed_separator(pdf)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(USABLE_WIDTH / 2, 6, "Grand Total")
    pdf.cell(USABLE_WIDTH / 2, 6, f"{grand_total:.2f}", align="R", new_x="LMARGIN", new_y="NEXT")

    content_height_mm = min(pdf.get_y() + MARGIN, page_height)

    pdf.output(output_path)
    trim_to_content(output_path, page_height, content_height_mm)
    return grand_total


def trim_to_content(pdf_path, page_height_mm, content_height_mm):
    doc = fitz.open(pdf_path)
    page = doc[0]
    width_pt = page.rect.width
    top_pt = page_height_mm * MM_TO_PT
    bottom_pt = (page_height_mm - content_height_mm) * MM_TO_PT
    page.set_mediabox(fitz.Rect(0, bottom_pt, width_pt, top_pt))
    doc.saveIncr()
    doc.close()


def log_transaction(items, delivery_charge, grand_total, bill_name):
    if os.path.exists(EXCEL_PATH):
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Transactions"
        sheet.append(EXCEL_HEADERS)

    bill_date = date.today().isoformat()
    for item in items:
        subtotal = item["price"] * item["quantity"]
        sheet.append(
            [bill_date, bill_name, item["name"], item["quantity"], item["price"], subtotal, delivery_charge, grand_total]
        )

    workbook.save(EXCEL_PATH)


def show_sales_summary():
    if not os.path.exists(EXCEL_PATH):
        print("No transactions recorded yet.")
        return

    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active

    bills = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        bill_date, bill_name, _product, _qty, _price, _subtotal, _delivery, grand_total = row
        if bill_name not in bills:
            bills[bill_name] = (bill_date, grand_total)

    if not bills:
        print("No transactions recorded yet.")
        return

    today = date.today().isoformat()
    this_month = date.today().strftime("%Y-%m")

    def totals_for(predicate):
        matching = [gt for d, gt in bills.values() if predicate(d)]
        return len(matching), sum(matching)

    today_count, today_total = totals_for(lambda d: d == today)
    month_count, month_total = totals_for(lambda d: d.startswith(this_month))
    all_count, all_total = totals_for(lambda d: True)

    print("\n--- Sales Summary ---")
    print(f"Today ({today}): {today_count} bill(s), Total: {today_total:.2f}")
    print(f"This Month ({this_month}): {month_count} bill(s), Total: {month_total:.2f}")
    print(f"All Time: {all_count} bill(s), Total: {all_total:.2f}")
    print("---------------------\n")


def generate_bill():
    items = get_items()
    if not items:
        print("No items entered. Nothing to print.")
        return

    delivery_charge = get_delivery_charge()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"bill_{date.today().isoformat()}_{len(os.listdir(OUTPUT_DIR)) + 1}.pdf"
    output_path = os.path.join(OUTPUT_DIR, filename)

    grand_total = build_pdf(items, delivery_charge, output_path)
    log_transaction(items, delivery_charge, grand_total, filename)

    print(f"\nBill saved to: {output_path}")
    print(f"Transaction logged to: {EXCEL_PATH}")
    print(f"Grand Total: {grand_total:.2f}")


def main():
    print("1. Generate a new bill")
    print("2. View sales summary")
    choice = input("Choose an option (1/2): ").strip()

    if choice == "2":
        show_sales_summary()
        return

    generate_bill()


if __name__ == "__main__":
    main()
